# backend/tests/api/test_export.py
"""
Tests de integración de los endpoints de export (CSV + XLSX + manifest).
Requieren Neon activo — usan el pool real vía app.state.pool.
"""
from io import BytesIO

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook

from api.database import create_pool
from api.main import app

BASE = "http://test"
SOURCES = ["cead", "dpp", "pdi", "ine_poblacion", "fiscalia"]


@pytest_asyncio.fixture
async def client():
    app.state.pool = await create_pool()
    try:
        async with AsyncClient(transport=ASGITransport(app=app), base_url=BASE) as c:
            yield c
    finally:
        await app.state.pool.close()


# ---------------------------------------------------------------------------
# Manifest
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_manifest_estructura(client):
    r = await client.get("/api/v1/export/manifest")
    assert r.status_code == 200
    data = r.json()

    assert "snapshot_utc" in data
    assert "fuentes" in data
    assert len(data["fuentes"]) == 5

    ids = {f["id"] for f in data["fuentes"]}
    assert ids == set(SOURCES)

    for f in data["fuentes"]:
        for key in (
            "titulo",
            "organismo",
            "fuente_url",
            "granularidad",
            "columnas",
            "caveat",
            "filas",
            "rango_anios",
            "descargas",
        ):
            assert key in f, f"Falta {key} en {f['id']}"
        assert f["fuente_url"].startswith("http"), "URL oficial debe ser http(s)"
        assert f["descargas"]["csv"].endswith(".csv")
        assert f["descargas"]["xlsx"].endswith(".xlsx")
        assert f["filas"] >= 0
        assert len(f["columnas"]) > 0


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("source", SOURCES)
@pytest.mark.asyncio
async def test_csv_headers_y_contenido(client, source):
    r = await client.get(f"/api/v1/export/{source}.csv")
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("text/csv")
    assert "attachment" in r.headers["content-disposition"]
    assert "filename=" in r.headers["content-disposition"]
    assert r.headers["x-fuente-oficial"].startswith("http")

    body = r.text
    lines = body.splitlines()
    # Debe tener al menos: 6 comentarios + 1 header + 1+ filas
    assert len(lines) > 7, f"{source}: CSV muy corto ({len(lines)} líneas)"
    # Primeras líneas con `#`
    assert lines[0].startswith("# Observatorio")
    assert any(l.startswith("# Fuente oficial:") for l in lines[:6])
    # Header = primera línea sin `#`
    header_line = next(l for l in lines if not l.startswith("#"))
    assert "anio" in header_line


@pytest.mark.asyncio
async def test_csv_fuente_desconocida(client):
    r = await client.get("/api/v1/export/foobar.csv")
    assert r.status_code == 404


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("source", SOURCES)
@pytest.mark.asyncio
async def test_xlsx_estructura(client, source):
    r = await client.get(f"/api/v1/export/{source}.xlsx")
    assert r.status_code == 200
    assert (
        r.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in r.headers["content-disposition"]

    wb = load_workbook(BytesIO(r.content), read_only=True)
    assert "datos" in wb.sheetnames
    assert "metadata" in wb.sheetnames

    # Metadata: primera columna son claves esperadas
    ws_meta = wb["metadata"]
    rows = list(ws_meta.iter_rows(values_only=True))
    keys = {row[0] for row in rows if row[0]}
    assert {
        "titulo",
        "organismo",
        "fuente_oficial_url",
        "granularidad",
        "snapshot_utc",
        "caveat_metodologico",
    }.issubset(keys)

    # Datos: al menos header + 1 fila
    ws_data = wb["datos"]
    data_rows = list(ws_data.iter_rows(values_only=True))
    assert len(data_rows) >= 2
    assert "anio" in data_rows[0]
