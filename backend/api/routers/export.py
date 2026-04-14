# backend/api/routers/export.py
#
# Exporta las tres bases de datos del Observatorio (CEAD, DPP, PDI) en CSV y XLSX
# como streaming responses. El objetivo es que cualquier ciudadano pueda descargar
# el dato crudo sin intermediarios, con referencia explícita a la fuente oficial.
#
# Diseño:
#   - CSV: asyncpg cursor async + StreamingResponse → memoria constante aunque
#     cead_hechos tenga 50k+ filas. Primera fila es un comentario con la URL de
#     la fuente oficial y la fecha del snapshot.
#   - XLSX: openpyxl en write_only mode → filas se van escribiendo a un buffer sin
#     cargar todo el workbook a memoria. Dos hojas: "datos" y "metadata".
#   - Manifest: resumen JSON (conteos, rangos de años) para que el frontend pinte
#     la página /datos sin hard-coding.

import csv
import io
from datetime import datetime, timezone
from typing import AsyncGenerator

import asyncpg
from fastapi import APIRouter, Depends, Query
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook

from api.database import get_pool

router = APIRouter(prefix="/api/v1/export", tags=["export"])

# Rango temporal aceptado por el observatorio para descargas: 2005-2025 inclusive.
# CEAD arranca en 2005; INE proyecta hasta 2035 pero recortamos para coherencia con
# el resto de las fuentes y para evitar publicar proyecciones futuras como dato.
EXPORT_ANIO_MIN = 2005
EXPORT_ANIO_MAX = 2025


# ---------------------------------------------------------------------------
# Metadata de fuentes (única fuente de verdad para CSV/XLSX/manifest)
# ---------------------------------------------------------------------------

SOURCES: dict[str, dict] = {
    "ine_poblacion": {
        "table": "ine_poblacion",
        "titulo": "Proyecciones de Población Comunal INE 2002-2035",
        "organismo": "Instituto Nacional de Estadísticas",
        "fuente_url": (
            "https://www.ine.gob.cl/estadisticas-por-tema/demografia-y-poblacion"
            "/proyecciones-de-poblacion"
        ),
        "granularidad": "Comunal (346 comunas)",
        "descripcion": (
            "Población proyectada por comuna para toda la serie 2002–2035. "
            "Denominador oficial para calcular tasas de detenciones por 100.000 habitantes."
        ),
        "etiqueta_eje": "Contexto",
        "vista_analitica": None,
        "caveat": (
            "Base Censo 2017. Actualización con base Censo 2024 esperada 2026. "
            "CUT sin zero-padding (ej: '1101', no '01101')."
        ),
        "columnas": ["cut", "nombre", "anio", "poblacion_total"],
        "sql": (
            "SELECT cut, nombre, anio, poblacion_total "
            "FROM ine_poblacion "
            "ORDER BY cut, anio"
        ),
    },
    "cead": {
        "table": "cead_hechos",
        "titulo": "CEAD · Casos policiales Ley 20.000",
        "organismo": "Subsecretaría de Prevención del Delito",
        "fuente_url": "https://cead.minsegpublica.gob.cl/estadisticas-delictuales/",
        "granularidad": "Comunal (344 de 346 comunas)",
        "descripcion": (
            "Casos policiales y personas aprehendidas por Carabineros + PDI en 7 "
            "subgrupos de la Ley 20.000 (tráfico, microtráfico, consumo vía pública, etc.). "
            "Frecuencia absoluta y tasa por 100k habitantes."
        ),
        "etiqueta_eje": "Policial",
        "vista_analitica": "/",
        "columnas": [
            "anio",
            "comuna_id",
            "subgrupo_id",
            "tipo_caso",
            "taxonomy_version",
            "frecuencia",
            "tasa_100k",
        ],
        "sql": (
            "SELECT anio, comuna_id, subgrupo_id, tipo_caso, taxonomy_version, "
            "frecuencia, tasa_100k FROM cead_hechos "
            "ORDER BY anio DESC, comuna_id, subgrupo_id"
        ),
        "caveat": (
            "Taxonomía delictiva cambia en 2024: 2005-2023 usan DMCS_legacy, "
            "2024+ usan Taxonomia_2024. Series temporales que crucen 2023-2024 "
            "NO son directamente comparables."
        ),
    },
    "dpp": {
        "table": "dpp_causas",
        "titulo": "DPP · Causas de la Defensoría Penal Pública",
        "organismo": "Defensoría Penal Pública",
        "fuente_url": "https://www.dpp.cl/pag/116/627/estadisticas",
        "granularidad": "Nacional + Regional (16 regiones)",
        "descripcion": (
            "Ingresos y términos de causas Ley de Drogas atendidas por la Defensoría. "
            "Incluye formas de término: absolución, condena, salida alternativa, "
            "derivación, facultativos y otras."
        ),
        "etiqueta_eje": "Defensorial",
        "vista_analitica": "/dpp",
        "columnas": [
            "anio",
            "region_id",
            "region_nombre",
            "tipo_delito",
            "tipo_termino",
            "n_causas",
            "n_imputados",
            "fuente",
        ],
        "sql": (
            "SELECT anio, region_id, region_nombre, tipo_delito, tipo_termino, "
            "n_causas, n_imputados, fuente FROM dpp_causas "
            "ORDER BY anio DESC, region_id NULLS FIRST, tipo_delito"
        ),
        "caveat": (
            "Sin granularidad comunal. Solo datos nacionales (comuna_id NULL) y "
            "regionales. Cobertura 2020-2024."
        ),
    },
    "fiscalia": {
        "table": "fiscalia_boletin",
        "titulo": "Fiscalía Nacional · Boletines Estadísticos (Ley 20.000)",
        "organismo": "Fiscalía Nacional · Ministerio Público",
        "fuente_url": "https://www.fiscaliadechile.cl/persecucion-penal/estadisticas",
        "granularidad": "Regional (16 regiones)",
        "descripcion": (
            "Causas ingresadas y terminadas por el Ministerio Público en la "
            "categoría oficial 'Delitos Ley de Drogas'. Suma IC (imputado conocido) "
            "y ID (imputado desconocido) de los boletines anuales."
        ),
        "etiqueta_eje": "Persecución penal",
        "vista_analitica": "/fiscalia",
        "columnas": [
            "anio",
            "region_id",
            "region_nombre",
            "categoria_delito",
            "tipo_metrica",
            "n_causas",
            "fuente",
        ],
        "sql": (
            "SELECT anio, region_id, region_nombre, categoria_delito, "
            "tipo_metrica, n_causas, fuente FROM fiscalia_boletin "
            "ORDER BY anio DESC, region_id, tipo_metrica"
        ),
        "caveat": (
            "Ingresos y términos por región en la categoría oficial 'Delitos Ley "
            "de Drogas' (incluye Ley 20.000). Sin granularidad comunal. Las 4 "
            "fiscalías zonales de la Región Metropolitana se agregan como RM (13)."
        ),
    },
    "pdi": {
        "table": "datosgob_seguridad",
        "titulo": "PDI · datos.gob.cl · Delitos investigados",
        "organismo": "Policía de Investigaciones de Chile (publicado vía datos.gob.cl)",
        "fuente_url": (
            "https://datos.gob.cl/dataset/"
            "5373dac4-a77a-48b2-9a8b-9cef7311f941"
        ),
        "granularidad": "Regional (16 regiones)",
        "descripcion": (
            "Tres subcategorías oficiales publicadas en datos.gob.cl: delitos "
            "investigados, denuncias art. 18 y víctimas art. 18. Filtradas por "
            "keywords de drogas (Ley 20.000 / tráfico / microtráfico / porte / consumo)."
        ),
        "etiqueta_eje": "Investigativa",
        "vista_analitica": "/pdi",
        "columnas": [
            "anio",
            "region_id",
            "categoria",
            "subcategoria",
            "frecuencia",
            "fuente",
        ],
        "sql": (
            "SELECT anio, region_id, categoria, subcategoria, frecuencia, fuente "
            "FROM datosgob_seguridad "
            "ORDER BY anio DESC, region_id, categoria"
        ),
        "caveat": (
            "Filtrado por keywords — frágil ante cambios de nomenclatura del "
            "publicador. Reemisión anual vía CKAN API del portal datos.gob.cl."
        ),
    },
}


def _snapshot_ts() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _filename(source: str, ext: str, anio: int | None = None) -> str:
    base = f"observatorio_{SOURCES[source]['table']}"
    if anio is not None:
        return f"{base}_{anio}.{ext}"
    return f"{base}_v1.{ext}"


def _sql_with_year_filter(meta: dict, anio: int | None) -> tuple[str, list]:
    """
    Devuelve (sql, args).

    - Si `anio` es int, filtra exactamente ese año (validado en [EXPORT_ANIO_MIN,
      EXPORT_ANIO_MAX] por el Query del endpoint).
    - Si `anio` es None, **siempre** clampa al rango global del observatorio.
      Esto evita publicar registros fuera del rango temporal declarado en /datos
      (ej. INE proyecciones >2025, datos CEAD pre-2005 si los hubiera).
    """
    inner = meta["sql"]
    cols = ", ".join(meta["columnas"])
    if anio is None:
        wrapped = (
            f"SELECT {cols} FROM ({inner}) AS t "
            f"WHERE anio BETWEEN $1 AND $2"
        )
        return wrapped, [EXPORT_ANIO_MIN, EXPORT_ANIO_MAX]
    wrapped = f"SELECT {cols} FROM ({inner}) AS t WHERE anio = $1"
    return wrapped, [anio]


# ---------------------------------------------------------------------------
# CSV streaming
# ---------------------------------------------------------------------------

async def _stream_csv(
    source: str,
    pool: asyncpg.Pool,
    anio: int | None,
) -> AsyncGenerator[bytes, None]:
    meta = SOURCES[source]
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")

    def _flush() -> bytes:
        data = buf.getvalue().encode("utf-8")
        buf.seek(0)
        buf.truncate(0)
        return data

    # Comentarios de procedencia (lectores como pandas admiten skip via `comment='#'`).
    buf.write(f"# Observatorio Judicial Cannabico — {meta['titulo']}\n")
    buf.write(f"# Fuente oficial: {meta['fuente_url']}\n")
    buf.write(f"# Organismo: {meta['organismo']}\n")
    buf.write(f"# Granularidad: {meta['granularidad']}\n")
    buf.write(f"# Snapshot: {_snapshot_ts()}\n")
    if anio is not None:
        buf.write(f"# Filtro: anio = {anio}\n")
    buf.write(f"# Caveat: {meta['caveat']}\n")
    yield _flush()

    writer.writerow(meta["columnas"])
    yield _flush()

    sql, args = _sql_with_year_filter(meta, anio)

    # BOM no — usamos UTF-8 puro; Excel moderno lo abre bien y evita duplicar bytes.
    async with pool.acquire() as conn:
        async with conn.transaction():
            async for row in conn.cursor(sql, *args):
                writer.writerow([row[c] if row[c] is not None else "" for c in meta["columnas"]])
                # Flush cada ~200 filas para no acumular.
                if buf.tell() > 32_768:
                    yield _flush()

    tail = _flush()
    if tail:
        yield tail


@router.get("/{source}.csv")
async def export_csv(
    source: str,
    anio: int | None = Query(None, ge=EXPORT_ANIO_MIN, le=EXPORT_ANIO_MAX, description=f"Filtrar por año específico ({EXPORT_ANIO_MIN}-{EXPORT_ANIO_MAX}). Si se omite, descarga el rango completo."),
    pool: asyncpg.Pool = Depends(get_pool),
):
    if source not in SOURCES:
        return JSONResponse({"error": f"Fuente desconocida: {source}"}, status_code=404)

    return StreamingResponse(
        _stream_csv(source, pool, anio),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{_filename(source, "csv", anio)}"',
            "X-Fuente-Oficial": SOURCES[source]["fuente_url"],
            "Cache-Control": "public, max-age=3600",
        },
    )


# ---------------------------------------------------------------------------
# XLSX streaming (write_only mode)
# ---------------------------------------------------------------------------

async def _build_xlsx(source: str, pool: asyncpg.Pool, anio: int | None) -> bytes:
    meta = SOURCES[source]
    wb = Workbook(write_only=True)

    # Hoja "datos"
    ws = wb.create_sheet("datos")
    ws.append(meta["columnas"])

    sql, args = _sql_with_year_filter(meta, anio)

    async with pool.acquire() as conn:
        async with conn.transaction():
            async for row in conn.cursor(sql, *args):
                ws.append([row[c] for c in meta["columnas"]])

    # Hoja "metadata"
    ws_meta = wb.create_sheet("metadata")
    ws_meta.append(["campo", "valor"])
    ws_meta.append(["titulo", meta["titulo"]])
    ws_meta.append(["organismo", meta["organismo"]])
    ws_meta.append(["fuente_oficial_url", meta["fuente_url"]])
    ws_meta.append(["granularidad", meta["granularidad"]])
    ws_meta.append(["snapshot_utc", _snapshot_ts()])
    if anio is not None:
        ws_meta.append(["filtro_anio", str(anio)])
    ws_meta.append(["caveat_metodologico", meta["caveat"]])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/{source}.xlsx")
async def export_xlsx(
    source: str,
    anio: int | None = Query(None, ge=EXPORT_ANIO_MIN, le=EXPORT_ANIO_MAX, description=f"Filtrar por año específico ({EXPORT_ANIO_MIN}-{EXPORT_ANIO_MAX}). Si se omite, descarga el rango completo."),
    pool: asyncpg.Pool = Depends(get_pool),
):
    if source not in SOURCES:
        return JSONResponse({"error": f"Fuente desconocida: {source}"}, status_code=404)

    payload = await _build_xlsx(source, pool, anio)
    return StreamingResponse(
        io.BytesIO(payload),
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{_filename(source, "xlsx", anio)}"',
            "X-Fuente-Oficial": SOURCES[source]["fuente_url"],
            "Content-Length": str(len(payload)),
            "Cache-Control": "public, max-age=3600",
        },
    )


# ---------------------------------------------------------------------------
# Manifest (metadata + conteos reales desde BD)
# ---------------------------------------------------------------------------

@router.get("/manifest")
async def export_manifest(pool: asyncpg.Pool = Depends(get_pool)):
    """
    Resumen dinámico de cada fuente: filas reales, rango de años, columnas,
    URL oficial, caveat metodológico. Consumido por /datos en el frontend.
    """
    items: list[dict] = []
    async with pool.acquire() as conn:
        for source_id, meta in SOURCES.items():
            # Skip tolerante: si la tabla aún no existe (ETL no ejecutado todavía),
            # la fuente no aparece en el manifest en vez de reventar el endpoint.
            try:
                # Conteos clampeados al rango global [EXPORT_ANIO_MIN, EXPORT_ANIO_MAX].
                row = await conn.fetchrow(
                    f"SELECT COUNT(*)::int AS filas, "
                    f"       MIN(anio)::int AS min_anio, "
                    f"       MAX(anio)::int AS max_anio "
                    f"FROM {meta['table']} "
                    f"WHERE anio BETWEEN $1 AND $2",
                    EXPORT_ANIO_MIN, EXPORT_ANIO_MAX,
                )
                # Lista exacta de años con registros dentro del rango — alimenta los chips.
                anios_rows = await conn.fetch(
                    f"SELECT DISTINCT anio FROM {meta['table']} "
                    f"WHERE anio BETWEEN $1 AND $2 "
                    f"ORDER BY anio",
                    EXPORT_ANIO_MIN, EXPORT_ANIO_MAX,
                )
            except asyncpg.exceptions.UndefinedTableError:
                continue
            anios_disponibles = [int(r["anio"]) for r in anios_rows]
            items.append({
                "id": source_id,
                "titulo": meta["titulo"],
                "organismo": meta["organismo"],
                "fuente_url": meta["fuente_url"],
                "granularidad": meta["granularidad"],
                "descripcion": meta.get("descripcion"),
                "etiqueta_eje": meta.get("etiqueta_eje"),
                "vista_analitica": meta.get("vista_analitica"),
                "columnas": meta["columnas"],
                "caveat": meta["caveat"],
                "filas": int(row["filas"] or 0),
                "rango_anios": {
                    "min": row["min_anio"],
                    "max": row["max_anio"],
                },
                "anios_disponibles": anios_disponibles,
                "descargas": {
                    "csv": f"/api/v1/export/{source_id}.csv",
                    "xlsx": f"/api/v1/export/{source_id}.xlsx",
                },
            })

    return {
        "snapshot_utc": _snapshot_ts(),
        "fuentes": items,
    }
